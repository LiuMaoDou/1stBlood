import os
import json
import socket
import struct
import time
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from datetime import datetime


def input_file_path():
    window = tk.Tk()
    window.title('(^_^))')
    window.geometry('220x20')

    l = tk.Label(window, text='"Pls Select The File"')
    l.pack()

    inpath = filedialog.askopenfilename(parent=window,
                                        initialdir=os.getcwd(),
                                        title="Please select a file:",
                                        filetypes=[('all files', '.*')])

    window.destroy()
    return inpath


class NetworkServer:
    SERVER_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'SERVER')
    if not os.path.exists(SERVER_FILE_PATH):
        os.makedirs(SERVER_FILE_PATH)

    def __init__(self, ip, port):
        self.ip = ip
        self.port = port

    def server_start(self):
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        sock.bind((self.ip, self.port))
        sock.listen(5)
        print("--> 服务已经启动...")

        while True:
            print("--> 等待客户端连接...")
            conn, addr = sock.accept()
            print(f"--> 连接成功, 连接信息: {addr}")

            while True:
                message = self.__recv_data(conn).decode('utf-8')

                if message == 'close':
                    print("--> 客户端断开连接...")
                    break

                message_type, content = message.split("|")
                content = content.replace('\'', '\"')
                content = json.loads(content)
                print(message_type, content)

                if message_type == 'register':
                    result = self.register(content)
                    if result == False:
                        print('--> 用户名已经存在')
                        self.__send_data(conn, '--> 用户名已经存在,无需注册...')
                    else:
                        self.__send_data(conn, '--> 用户注册成功...')
                        print('--> 用户注册成功...')

                if message_type == 'login':
                    result = self.login(content)
                    self.__send_data(conn, result)

                    if result == '用户认证失败':
                        continue

                    while True:
                        ls_result = self.__recv_data(conn).decode('utf-8')
                        if ls_result == 'ls' and os.path.exists(os.path.join(self.SERVER_FILE_PATH, content['username'])):
                            ls_result_back = '\n'.join(os.listdir(os.path.join(self.SERVER_FILE_PATH, content['username'])))
                            self.__send_data(conn, ls_result_back)

                        elif ls_result == 'upload_file':
                            upload_file_name = self.__recv_data(conn).decode('utf-8')
                            self.__recv_file(conn, os.path.join(content['username'], upload_file_name))

                        elif ls_result == 'download_file':
                            download_file_name = self.__recv_data(conn).decode('utf-8')
                            file_size, download_file_name = download_file_name.split('|')
                            download_file_name = os.path.join(self.SERVER_FILE_PATH, content['username'], download_file_name)
                            file_size = int(file_size)
                            print(os.stat(download_file_name).st_size)
                            print(file_size)
                            if os.stat(download_file_name).st_size == file_size:
                                self.__send_data(conn, '文件一样大, 不需要续传')
                                print('--> 文件一样大, 不需要续传')

                            else:
                                self.__send_data(conn, '开始续传')
                                print('开始续传')
                                self.__upload_file(conn, int(file_size), download_file_name)

                        elif ls_result == "Q":
                            break

            conn.close()
        sock.close()

    @staticmethod
    def __recv_data(conn, chunk_size=1024):
        has_read_size = 0
        bytes_list = []

        while has_read_size < 4:
            chunk = conn.recv(4 - has_read_size)
            has_read_size += len(chunk)
            bytes_list.append(chunk)
        header = b"".join(bytes_list)
        data_length = struct.unpack('i', header)[0]

        # 获取数据
        data_list = []
        has_read_data_size = 0

        while has_read_data_size < data_length:
            size = chunk_size if (data_length - has_read_data_size) > chunk_size else data_length - has_read_data_size
            chunk = conn.recv(size)
            data_list.append(chunk)
            has_read_data_size += len(chunk)

        data = b"".join(data_list)
        return data


    def __recv_file(self, conn, save_file_name, chunk_size=1024):
        save_file_path = os.path.join(self.SERVER_FILE_PATH, save_file_name)
        # 获取头部信息：数据长度
        has_read_size = 0
        bytes_list = []
        while has_read_size < 4:
            chunk = conn.recv(4 - has_read_size)
            bytes_list.append(chunk)
            has_read_size += len(chunk)
        header = b"".join(bytes_list)
        data_length = struct.unpack('i', header)[0]

        # 获取数据
        file_object = open(save_file_path, mode='wb')
        has_read_data_size = 0
        while has_read_data_size < data_length:
            size = chunk_size if (data_length - has_read_data_size) > chunk_size else data_length - has_read_data_size
            chunk = conn.recv(size)
            file_object.write(chunk)
            file_object.flush()
            has_read_data_size += len(chunk)

            time.sleep(0.0001)
            percent = round((int(has_read_data_size) / int(data_length)) * 100, 3)
            percent_bar = "▋" * int(round(percent / 10, 0))
            print("\r文件总大小为：{}字节，已下载{}字节, 进度{}%: {}".format(data_length, has_read_data_size, percent, percent_bar),
                  end="")
        file_object.close()

    @staticmethod
    def __send_data(conn, content):
        data = content.encode('utf-8')
        header = struct.pack('i', len(data))
        conn.sendall(header)
        conn.sendall(data)


    def register(self, dic):
        file_name = os.path.join(self.SERVER_FILE_PATH, 'user_infor_record.xlsx')
        border = Border(top=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"),
                        left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"))

        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "用户记录"
            ws['A1'] = '用户名'
            ws['A1'].alignment = Alignment(horizontal='center')
            ws['A1'].fill = PatternFill("solid", fgColor="3366cc")
            ws['A1'].font = Font(color="ffffff")
            ws['A1'].border = border
            ws['B1'] = '密码'
            ws['B1'].alignment = Alignment(horizontal='center')
            ws['B1'].fill = PatternFill("solid", fgColor="3366cc")
            ws['B1'].font = Font(color="ffffff")
            ws['B1'].border = border
            ws['C1'] = '注册时间'
            ws['C1'].alignment = Alignment(horizontal='center')
            ws['C1'].fill = PatternFill("solid", fgColor="3366cc")
            ws['C1'].font = Font(color="ffffff")
            ws['C1'].border = border
            wb.save(file_name)

        wb = load_workbook(file_name)
        ws = wb['用户记录']

        # 判断用户之前是否注册, 如果注册停止程序并且返回信息提示客户端
        for row in ws.rows:
            if dic['username'] == row[0].value:
                return False

        row = (len(list(ws.rows)) + 1)
        ws.cell(row=row, column=1).value = dic['username']
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).value = dic['pwd']
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=3).value = datetime.now()
        ws.cell(row=row, column=3).border = border

        os.makedirs(os.path.join(self.SERVER_FILE_PATH, dic['username']))
        wb.save(file_name)


    def login(self, dic):
        file_name = os.path.join(self.SERVER_FILE_PATH, 'user_infor_record.xlsx')
        if not os.path.exists(file_name):
            print('--> 没有用户注册信息...')
            return None

        wb = load_workbook(file_name)
        ws = wb['用户记录']

        for row in ws.rows:
            if dic['username'] == row[0].value and dic['pwd'] == row[1].value:
                print("--> 用户认证成功...")
                return '用户认证成功'

        print("--> 用户认证失败...")
        return '用户认证失败'


    def __upload_file(self, conn, file_size, file_path):
        print("--------------------------------")
        self.__send_data_file(conn, file_size, file_path)
        print("\n--------------------------------")
        print("--> 发送文件完成...")



    @staticmethod
    def __send_data_file(conn, xc_file_size, file_path):
        file_size = os.stat(file_path).st_size
        header = struct.pack('i', file_size-xc_file_size)
        conn.sendall(header)


        file_object = open(file_path, mode='rb')

        file_object.seek(xc_file_size)
        has_send_size = xc_file_size

        while has_send_size < file_size:
            chunk = file_object.read(2048)
            conn.sendall(chunk)
            has_send_size += len(chunk)

            time.sleep(0.0001)
            percent = round((int(has_send_size) / int(file_size)) * 100, 3)
            percent_bar = "▋" * int(round(percent / 10, 0))
            print("\r文件总大小为：{}字节，已发送{}字节, 进度{}%: {}".format(file_size, has_send_size, percent, percent_bar), end="")

        file_object.close()


if __name__ == "__main__":
    # IP和端口可以根据实际情况输入
    server = NetworkServer('127.0.0.1', 50000)
    server.server_start()
